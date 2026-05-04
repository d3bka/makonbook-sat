from pathlib import Path
import os
from django.contrib.auth import get_user_model
from apps.sat.models import Classroom

User = get_user_model()

desktop_candidates = [
    Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop",
    Path.home() / "OneDrive" / "Desktop",
    Path.home() / "Desktop",
    Path.cwd(),
]

outdir = next((p for p in desktop_candidates if p.exists()), Path.cwd())
xlsx_path = outdir / "makonbook_teachers_report.xlsx"
csv_path = outdir / "makonbook_teachers_report.csv"

teacher_ids = (
    Classroom.objects
    .exclude(teacher_id=None)
    .values_list("teacher_id", flat=True)
    .distinct()
)

teachers = User.objects.filter(id__in=teacher_ids).order_by(
    "first_name", "last_name", "username"
)

rows = []

for teacher in teachers:
    classrooms = Classroom.objects.filter(teacher_id=teacher.id)
    active_classrooms = classrooms.filter(is_active=True)

    approved_students = (
        classrooms
        .filter(memberships__role="student", memberships__status="approved")
        .values("memberships__user_id")
        .distinct()
        .count()
    )

    first_name = teacher.first_name or ""
    last_name = teacher.last_name or ""
    full_name = teacher.get_full_name().strip() or f"{first_name} {last_name}".strip()

    rows.append({
        "ID": teacher.id,
        "Full name": full_name,
        "First name": first_name,
        "Last name": last_name,
        "Username": teacher.username or "",
        "Email": teacher.email or "",
        "Total classrooms": classrooms.count(),
        "Active classrooms": active_classrooms.count(),
        "Approved students": approved_students,
    })

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Teachers"

    headers = [
        "ID",
        "Full name",
        "First name",
        "Last name",
        "Username",
        "Email",
        "Total classrooms",
        "Active classrooms",
        "Approved students",
    ]

    ws.append(["Makonbook Teachers Report"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row in rows:
        ws.append([row[h] for h in headers])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{ws.max_row}"

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Teachers with classrooms", len(rows)])
    summary.append(["Teachers with active classrooms", sum(1 for r in rows if r["Active classrooms"] > 0)])
    summary.append(["Total classrooms", sum(r["Total classrooms"] for r in rows)])
    summary.append(["Total active classrooms", sum(r["Active classrooms"] for r in rows)])
    summary.append(["Total approved students", sum(r["Approved students"] for r in rows)])

    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, max_col=2):
        for cell in row:
            cell.border = border

    summary.column_dimensions["A"].width = 35
    summary.column_dimensions["B"].width = 18

    wb.save(xlsx_path)
    print(f"Excel report created: {xlsx_path}")

except ImportError:
    import csv

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "ID",
            "Full name",
            "First name",
            "Last name",
            "Username",
            "Email",
            "Total classrooms",
            "Active classrooms",
            "Approved students",
        ])
        writer.writeheader()
        writer.writerows(rows)

    print("openpyxl is not installed, so CSV was created instead.")
    print(f"CSV report created: {csv_path}")

print(f"Total teachers with classrooms: {len(rows)}")
